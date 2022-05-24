import openpyxl as op
import xlwings as xw

import Mappings


def automationRunner(rawDataPath, putIntoPath, startRow, stopRow):
    print("reached automationRunner")
    copyToWorkbook(putIntoPath)

    for i in range(startRow,stopRow+1):
        print("reached init")
        automateTest = Automator(i, submissionName=rawDataPath, capName=putIntoPath)
        automateTest.create_Page()
        automateTest.populate_page()


def copyToWorkbook(putIntoPath):
    wb1 = xw.Book('CapSheet.xlsx')
    wb2 = xw.Book(putIntoPath)

    ws1 = wb1.sheets[0]
    ws1.api.copy_worksheet(after_=wb2.sheets[0].api)
    wb2.save()
    wb2.app.quit()

class Automator:

    eventCount = 0
    eventList = {}

    moveCells_ProgrammingBool = False
    moveCells_ProgrammingBool2 = False
    moveCells_seriesProgrammingBool = False
    moveCells_TripsCC1Bool = False
    moveCells_TripsCC2Bool = False
    moveCells_tripsOtherBool = False


    def __init__(self, row, submissionName, capName):
        #print("starting init")
        self.submissions = op.load_workbook(submissionName)
        self.submissionSheet = self.submissions.active
        # os.chdir(os.path.expanduser(folder))
        #print("part 2")
        self.capName=capName
        self.capSheet = op.load_workbook(capName)
        self.target = None
        self.row = str(row)

    def create_Page(self):

        print("in create_Page")
        #newWb = op.Workbook()

        cap = self.capSheet.get_sheet_by_name('CapSheet')

        self.target = self.capSheet.copy_worksheet(cap)

        try:
            self.target.title = self.submissionSheet['M' + str(self.row)].value
        except:
            self.target.title = "PUT ORG.NAME IN HERE"

        self.target['A1'] = self.target['A1'].value + self.submissionSheet['M' + str(self.row)].value
        self.target['K1'] = self.target['K1'].value + ' ' + str(self.submissionSheet['N' + str(self.row)].value)
        print(self.target['K1'].value)

        self.target['B1'] = self.submissionSheet['P' + str(self.row)].value
        self.target['C1'] = self.submissionSheet['Q' + str(self.row)].value
        self.target['D1'] = self.submissionSheet['R' + str(self.row)].value
        self.target['E1'] = self.submissionSheet['S' + str(self.row)].value
        self.target['F1'] = self.submissionSheet['T' + str(self.row)].value
        self.target['F1'] = self.submissionSheet['O' + str(self.row)].value


    def populate_page(self):
        print("in populate page")
        print(self.capSheet.sheetnames)

        #print(sheet['A1'].value)

        #Programs: BR, CW, EC, EY, FU, HA, IF, JA, JW, LC, MH, ND
        eventListColumns = {'BX' : 'program', 'CZ':'programSeries', 'EC': 'tripCC', 'EX': 'tripOther',
                            'FS': 'program', 'GW': 'programSeries', 'HY': 'tripCC', 'IS': 'tripOther', 'JN': 'program',
                            'KS': 'programSeries', 'LU':'tripCC', 'MP':'tripOther'}

        eventListRow = {key + str(self.row): value for key, value in eventListColumns.items()}
        print(eventListRow)


        for eventCell, type in eventListRow.items():
            if self.submissionSheet[eventCell].value is not None:
                self.eventList.update({eventCell:type})
                self.eventCount += 1


        print(self.eventList)
        self.populate_orgMaintenance()


        #Still need to account for >2 events
        for eventCell, type in self.eventList.copy().items():
            if type == 'program':
                self.populate_programStandAlone(eventCell)
            elif type == 'programSeries':
                self.populate_programSeries(eventCell)
            elif type == 'tripCC':
                self.populate_tripsCC(eventCell)
            elif type == 'tripOther':
                self.populate_tripsOther(eventCell)

            self.eventCount -= 1
            del self.eventList[eventCell]


        self.capSheet.save(self.capName)


    def populate_orgMaintenance(self):

        print("in populate_orgMaintenance")
        sheet = self.target

        for capColumn,submissionCell in Mappings.moveCells_OrgMaintenance.items():
            if self.submissionSheet[submissionCell] is not None:
                sheet[capColumn] = self.submissionSheet[submissionCell + self.row].value
            else:
                continue

        print("finished orgMaintenance")

    #TODO: Still need to do Contracts and Rights
    def populate_programStandAlone(self,eventCell):

        print("in populate_ProgrammingStandAlone")

        if self.moveCells_ProgrammingBool is False:

            self.iteration(Mappings.moveCells_Programming, eventCell, self.row, self.target)
            self.moveCells_ProgrammingBool= True

        elif self.moveCells_ProgrammingBool2 is False:

           self.iteration(Mappings.moveCells_Programming2, eventCell, self.row, self.target)
           self.moveCells_ProgrammingBool2 = True



    def populate_programSeries(self, eventCell):
        print("in populateProgrammingSeries")

        if self.moveCells_seriesProgrammingBool is False:

            self.iteration(Mappings.moveCells_seriesProgramming, eventCell, self.row, self.target)
            self.moveCells_seriesProgrammingBool = True

    #Populates CC Trips
    def populate_tripsCC(self, eventCell):
        print("in populate_tripsCC")

        if self.moveCells_TripsCC1Bool is False:

            self.iteration(Mappings.moveCells_tripsCC1, eventCell, self.row, self.target)
            self.moveCells_TripsCC1Bool = True

        elif self.moveCells_TripsCC2Bool is False:

            self.iteration(Mappings.moveCells_TripsCC2, eventCell, self.row, self.target)
            self.moveCells_TripsCC2Bool = True

    #Populates Other Trips
    def populate_tripsOther(self, eventCell):
        print("in populate_tripsOther")

        if self.moveCells_tripsOtherBool is False:
            self.iteration(Mappings.moveCells_otherTrip, eventCell, self.row, self.target)
            self.moveCells_tripsOtherBool = True


    #Iterates through cells
    def iteration(self, moveCells, eventCell, row, sheet):

        firstKey = next(iter(moveCells))
        li = moveCells[firstKey]
        liRow = [item + str(self.row) for item in li]
        print(liRow)
        index = liRow.index(eventCell)

        for capColumn, submissionCell in moveCells.items():
            submissionCell = submissionCell[index]
            if isinstance(submissionCell, list):
                for sub in submissionCell:
                    term = sub + row
                    print(self.submissionSheet[term].value)
                    if self.submissionSheet[sub + row].value is not None:
                        sheet[capColumn] = f"{sheet[capColumn].value}  {self.submissionSheet[sub + row].value}"
                print (sheet[capColumn].value)


            elif self.submissionSheet[submissionCell + row].value is not None:
                term = self.submissionSheet[submissionCell + row].value
                if isinstance(sheet[capColumn].value, str):
                    sheet[capColumn] = f"{sheet[capColumn].value} {term}"
                else:
                    sheet[capColumn] = term


















'''

 def DUMP(self, eventCell, type):
        print("In Dump")
        updated_dict = {}
        letter = 'N'

        if type == 'program':
            updated_dict =  {'O2': ['BX', 'FS', 'JN'], 'O3': ['BY', 'FT', 'JO'], 'O4': ['BZ', 'FU', 'JP'], 'O5': ['CA', 'FV', 'JQ'],
                             'O6': ['CB', 'FW', 'JR'], 'O7':['CC', 'FX', 'JS'], 'O8': ['CE', 'FZ', 'JU'], 'O9' : ['CF','GA', 'JV'], 'O10':['CG','GB','JW'],
                             'O11':['CH', 'GC', 'JX'], 'O12':['CI', 'GD', 'JY'], 'O13':['CJ','GE','JZ'], 'O14':['CK','GF','KA'], 'O15':['CL','GG','KB'],
                             'O16':['CM', 'GI', 'KE'], 'O17':['CU', 'GQ', 'KM'], 'O18':['CN', 'GJ', 'KF'], 'O19':['CO', 'GK', 'KG'], 'O20':['CP', 'GL', 'KH'],
                            'O21':['CQ', 'GM', 'KI'], 'O22':['CR', 'GN', 'KJ'], 'O23':['CS', 'GO', 'KK'], 'O24':['CV', 'GR', 'KN'], 'O25':['CW', 'GS', 'KO'],
                            'O27':['CX', 'GT', 'KP']
                             }
            

        elif type == 'programSeries':
            for value in self.moveCells_seriesProgramming.values():
                updated_dict[variable + '3'] = value

        elif type == 'tripCC':
            for value in self.moveCells_tripsCC1.values():
                updated_dict[variable + '3'] = value

        elif type == 'tripOther':
            for value in self.moveCells_otherTrip.values():
                updated_dict[variable + '3'] = value



'''


'''
      if isinstance(self.submissionSheet[submissionCell + row].value, datetime.date):
                            item = datetime.datetime.strftime(self.submissionSheet[submissionCell + row].value, datetime.date)
                        sheet[capColumn] = sheet[capColumn].value + str(item)
                        
                        if sheet[capColumn].value is not None:
                        item = self.submissionSheet[submissionCell + row].value
                  

                    else: sheet[capColumn] + self.submissionSheet[submissionCell + row].value


'''